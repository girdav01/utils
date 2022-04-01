'''
David Girard Sr Product Manager Trend Micro
Get MITRE Evaluation #4 raw json results
March 31st 2022
tested with Python 3.9 but most 3.x should works
'''
import requests  # pip3 install requests
import json
from openpyxl import Workbook  # pip3 install openpyxl
from openpyxl.styles import Font, Color


base_url = "https://attackevals.mitre-engenuity.org/api/export/?participant="
suffix = "&adversary=wizard-spider-sandworm"  #each json file on mitre-engenuity.org got this
filesuffix = '_wizard-spider-sandworm.json'

#list of vendors in this year eval
vendors = ['TrendMicro','Bitdefender', 'CheckPoint', 'Cisco', 'CrowdStrike', 'Cybereason', 'CyCraft', 'Cylance',
           'Cynet', 'Deepinstinct', 'Elastic','ESET', 'Fidelis', 'FireEye', 'Fortinet','f-secure', 'Malwarebytes',
           'McAfee', 'Microsoft', 'PaloAltoNetworks', 'Qualys', 'Rapid7', 'ReaQta', 'SentinelOne', 'Somma',
           'Sophos', 'Symantec', 'AhnLab', 'Uptycs', 'VMware']

def getResults(v):
    print(v+suffix)
    r_file = requests.get(base_url + v + suffix, verify=False)
    json_data = json.loads(r_file.text)
    with open('data/' + v+filesuffix, 'w', encoding='utf-8') as f:
        json.dump(json_data, f, ensure_ascii=False, indent=4)

def downloadRawData():
    for vendor in vendors:
        getResults(vendor)

def buildSummary():
    # build a summary in an excel shet
    wb = Workbook()
    sh = wb.active
    # set the columns names
    sh['A1'] = "Participant"
    cols = ['Total_Substeps', 'Total_Detections',
            'Analytic_Detections', 'Telemetry_Detections','Analytic_Coverage','Telemetry_Coverage','Visibility']
    col_start = 2
    for i in range(0, len(cols)):
        sh.cell(1, col_start+i).value = cols[i]
    sh.cell(1, col_start + i +1).value = "Linux Capability"
    sh.cell(1, col_start + i + 2).value = "Protection Capability"
    sh.cell(1, col_start + i + 3).value = "Analytic Coverage %"
    sh.cell(1, col_start + i + 4).value = "Telemetry Coverage %"
    sh.cell(1, col_start + i + 5).value = "Visibility %"

    # now get the vendor summary data in the sheet
    row = 2
    red_text = Font(color="00FF0000")
    for vendor in vendors:
        print(vendor) # just to see it'S processing or crashing

        with open('data/' + vendor+filesuffix, encoding='utf-8') as f:
            data = json.load(f)
            icol = 1
            i = 1
            # Participant = vendor
            sh.cell(row, icol).value = data[0]['Participant_Name']
            # get the summary data under Aggregates
            for col in cols:
                sh.cell(row,icol + i).value = data[0]['Adversaries'][0]['Aggregate_Data']['Aggregates'][col]
                i=i+1
            # Add the capabilities.
            if 'Linux Capability' in data[0]['Adversaries'][0]['Participant_Capabilities']:
                sh.cell(row, icol + i).value = "Yes"
            else:
                sh.cell(row, icol + i).value = "No"
                sh.cell(row, icol + i).font = red_text

            if 'Protection Capability' in data[0]['Adversaries'][0]['Participant_Capabilities']:
                sh.cell(row, icol + i +1 ).value = "Yes"
            else:
                sh.cell(row, icol + i + 1).value = "No"
                sh.cell(row, icol + i + 1).font = red_text
            # Add the %
            percent_analytics = eval(data[0]['Adversaries'][0]['Aggregate_Data']['Aggregates']["Analytic_Coverage"]+ '*100')
            sh.cell(row, icol + i + 2).value = percent_analytics

            percent_telemetry = eval(data[0]['Adversaries'][0]['Aggregate_Data']['Aggregates']["Telemetry_Coverage"]+ '*100')
            sh.cell(row, icol + i + 3).value = percent_telemetry

            percent_visibility = eval(data[0]['Adversaries'][0]['Aggregate_Data']['Aggregates']["Visibility"]+ '*100')
            sh.cell(row, icol + i + 4).value = percent_visibility
            f.close()
        row=row +1

    # add your improvements, like creating charts

    # change the name to watever you like
    wb.save('data/wizard-spider-sandworm.xlsx')

# simple : download and build summary.
downloadRawData()
buildSummary()
# add your steps here... pandas maybe, or upload JSON's to Elastic

